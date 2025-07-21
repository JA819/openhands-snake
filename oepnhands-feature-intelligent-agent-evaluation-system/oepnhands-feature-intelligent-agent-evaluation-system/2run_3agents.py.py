import requests
import time
import os
from openpyxl import Workbook, load_workbook

# --------------------------
# 1. 三个 Agent 的配置
# --------------------------
AGENTS = {
    "bge-m3": {
        "url": "https://aigov-dify.bytebroad.com.cn/v1/chat-messages",
        "key": "Bearer app-0R77MRBexBD7iCn87aoC1X69"
    },
    "qwen": {
        "url": "https://aigov-dify.bytebroad.com.cn/v1/chat-messages",
        "key": "Bearer app-TAwq1XsLmXXLzYKTQTJV2nUe"
    },
    "rangflow": {
        "url": "https://aigov-dify.bytebroad.com.cn/v1/chat-messages",
        "key": "Bearer app-AeGMfmHfQl2pMN5xVKYIbGEG"
    }
}

# --------------------------
# 2. 文件路径
# --------------------------
GOLDEN_SET_FILE = r"C:/Users/19541/Desktop/Internship services/Internship services/Dify/learnproject/learn HTTP/问答对1.xlsx"
OUTPUT_FILE     = r"C:/Users/19541/Desktop/Internship services/Internship services/Dify/learnproject/learn HTTP/evaluation_results_3agents.xlsx"

# --------------------------
# 3. 通用函数
# --------------------------
def call_dify(agent_name, question, user_id="eval_user"):
    cfg = AGENTS[agent_name]
    payload = {
        "inputs": {},
        "query": question,
        "response_mode": "blocking",
        "user": user_id
    }
    headers = {
        "Authorization": cfg["key"],
        "Content-Type": "application/json"
    }
    try:
        resp = requests.post(cfg["url"], json=payload, headers=headers, timeout=60)
        resp.raise_for_status()
        return resp.json().get("answer", "")
    except Exception as e:
        return f"Error calling {agent_name}: {e}"

def read_questions(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    return [{"question": r[0], "answer": r[1]}
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[0] and r[1]]

def write_results(results_dict):
    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    for agent, rows in results_dict.items():
        ws = wb.create_sheet(title=agent)
        ws.append(["问题", "标准答案", f"{agent} 生成的答案"])
        for r in rows:
            ws.append([r["question"], r["standard_answer"], r["agent_answer"]])
    wb.save(OUTPUT_FILE)
    print(f"结果已保存至 {OUTPUT_FILE}")

# --------------------------
# 4. 主流程
# --------------------------
if __name__ == "__main__":
    questions = read_questions(GOLDEN_SET_FILE)
    results = {k: [] for k in AGENTS}

    for idx, item in enumerate(questions, 1):
        print(f"正在处理第 {idx}/{len(questions)} 个问题 ...")
        q = item["question"]
        for agent in AGENTS:
            ans = call_dify(agent, q, f"{agent}_user_{idx}")
            results[agent].append({
                "question": q,
                "standard_answer": item["answer"],
                "agent_answer": ans
            })
        time.sleep(1.5)   # 控制频率

    write_results(results)
    print("全部完成！")
