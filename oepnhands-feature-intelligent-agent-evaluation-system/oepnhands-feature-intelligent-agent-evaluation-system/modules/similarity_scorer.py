import aiohttp
import json
import os
import jieba
import numpy as np
from openpyxl import load_workbook
from sklearn.metrics.pairwise import cosine_similarity
from typing import Dict, List, Any
import asyncio

class SimilarityScorer:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.embed_url = "https://api.siliconflow.cn/v1/embeddings"
        self.headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }

    async def get_embedding(self, text: str, retries: int = 3) -> List[float]:
        """获取文本向量"""
        if not isinstance(text, str) or not text.strip():
            return []
        
        payload = {
            "model": "Qwen/Qwen3-Embedding-0.6B",
            "input": " ".join(jieba.lcut(text)),
            "encoding_format": "float"
        }
        
        for attempt in range(retries):
            try:
                async with aiohttp.ClientSession() as session:
                    async with session.post(self.embed_url, json=payload, headers=self.headers, timeout=30) as response:
                        response.raise_for_status()
                        result = await response.json()
                        embedding = result["data"][0]["embedding"]
                        return embedding
            except Exception as e:
                print(f"embedding 失败 (尝试 {attempt + 1}/{retries}): {e}")
                if attempt < retries - 1:
                    await asyncio.sleep(1)  # 等待1秒后重试
        
        return []

    async def calculate_similarity_scores(self, standard_text: str, generated_text: str) -> Dict[str, float]:
        """计算相似度评分"""
        # 获取向量
        emb_std = await self.get_embedding(standard_text)
        emb_gen = await self.get_embedding(generated_text)
        
        # 余弦相似度
        cosine_sim = 0.0
        if emb_std and emb_gen and len(emb_std) == len(emb_gen):
            cosine_sim = cosine_similarity([emb_std], [emb_gen])[0][0]

        # Jaccard相似度
        set_std = set(jieba.lcut(standard_text))
        set_gen = set(jieba.lcut(generated_text))
        jaccard_sim = len(set_std & set_gen) / len(set_std | set_gen) if set_std | set_gen else 0.0

        # 综合相似度评分
        weighted_score = 0.7 * cosine_sim + 0.3 * jaccard_sim

        return {
            "cosine_similarity": float(cosine_sim),
            "jaccard_similarity": float(jaccard_sim),
            "weighted_score": float(weighted_score)
        }

    async def calculate_scores(self, input_file_path: str, output_file_path: str) -> Dict[str, Any]:
        """计算Excel文件中所有答案的相似度评分"""
        wb = load_workbook(input_file_path)
        
        results = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 添加表头（如果不存在）
            if ws.cell(row=1, column=4).value != "余弦相似度":
                ws["D1"] = "余弦相似度"
                ws["E1"] = "Jaccard相似度"
                ws["F1"] = "综合相似度评分"
            
            scores = []
            row_count = 0
            
            # 从第2行开始遍历
            for row in ws.iter_rows(min_row=2, max_col=6, values_only=False):
                if row[1].value and row[2].value:  # 确保标准答案和生成答案都存在
                    standard_answer = str(row[1].value)
                    generated_answer = str(row[2].value)
                    
                    # 计算相似度
                    similarity = await self.calculate_similarity_scores(standard_answer, generated_answer)
                    
                    # 写入Excel
                    row[3].value = similarity["cosine_similarity"]
                    row[4].value = similarity["jaccard_similarity"]
                    row[5].value = similarity["weighted_score"]
                    
                    scores.append(similarity["weighted_score"])
                    row_count += 1
                    
                    # 添加延迟避免API限制
                    await asyncio.sleep(0.1)
            
            # 计算统计信息
            if scores:
                results[sheet_name] = {
                    "count": row_count,
                    "mean_score": float(np.mean(scores)),
                    "max_score": float(np.max(scores)),
                    "min_score": float(np.min(scores)),
                    "std_score": float(np.std(scores))
                }
            
            print(f"已处理 {sheet_name} sheet，共 {row_count} 行数据")
        
        # 保存结果
        wb.save(output_file_path)
        print(f"相似度评分结果已保存至 {output_file_path}")
        
        return results

    def analyze_scores(self, file_path: str) -> Dict[str, Any]:
        """分析已计算的相似度评分"""
        wb = load_workbook(file_path)
        analysis = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            scores = []
            
            # 读取综合相似度评分列（第6列）
            for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
                if row[5] is not None:  # 综合相似度评分
                    scores.append(float(row[5]))
            
            if scores:
                scores_array = np.array(scores)
                
                # 计算统计信息
                analysis[sheet_name] = {
                    "count": len(scores),
                    "mean_score": float(np.mean(scores_array)),
                    "median_score": float(np.median(scores_array)),
                    "max_score": float(np.max(scores_array)),
                    "min_score": float(np.min(scores_array)),
                    "std_score": float(np.std(scores_array)),
                    "high_score_ratio": float(np.sum(scores_array >= 0.8) / len(scores_array)),
                    "low_score_ratio": float(np.sum(scores_array < 0.5) / len(scores_array)),
                    "score_distribution": {
                        "excellent": int(np.sum(scores_array >= 0.9)),
                        "good": int(np.sum((scores_array >= 0.7) & (scores_array < 0.9))),
                        "fair": int(np.sum((scores_array >= 0.5) & (scores_array < 0.7))),
                        "poor": int(np.sum(scores_array < 0.5))
                    }
                }
        
        return analysis

    def get_detailed_results(self, file_path: str, sheet_name: str = None) -> List[Dict]:
        """获取详细的评分结果"""
        wb = load_workbook(file_path)
        
        if sheet_name and sheet_name in wb.sheetnames:
            sheets_to_process = [sheet_name]
        else:
            sheets_to_process = wb.sheetnames
        
        detailed_results = []
        
        for sheet in sheets_to_process:
            ws = wb[sheet]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:  # 确保问题、标准答案、生成答案都存在
                    result = {
                        "agent": sheet,
                        "question": str(row[0]),
                        "standard_answer": str(row[1]),
                        "generated_answer": str(row[2]),
                        "cosine_similarity": float(row[3]) if row[3] is not None else 0.0,
                        "jaccard_similarity": float(row[4]) if row[4] is not None else 0.0,
                        "weighted_score": float(row[5]) if row[5] is not None else 0.0
                    }
                    detailed_results.append(result)
        
        return detailed_results