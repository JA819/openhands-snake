import pandas as pd
import numpy as np
from openpyxl import load_workbook
from typing import Dict, List, Any
import json

class Analyzer:
    def __init__(self):
        pass

    async def analyze_file(self, file_path: str) -> Dict[str, Any]:
        """分析相似度评分文件"""
        try:
            wb = load_workbook(file_path)
            analysis_result = {
                "agents": {},
                "comparison": {},
                "best_agent": None,
                "overall_stats": {}
            }
            
            agent_scores = {}
            
            # 分析每个智能体的表现
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                scores = []
                questions_data = []
                
                # 读取数据
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 6 and row[5] is not None:  # 确保有综合相似度评分
                        score = float(row[5])
                        scores.append(score)
                        
                        questions_data.append({
                            "question": str(row[0]) if row[0] else "",
                            "standard_answer": str(row[1]) if row[1] else "",
                            "generated_answer": str(row[2]) if row[2] else "",
                            "cosine_similarity": float(row[3]) if row[3] is not None else 0.0,
                            "jaccard_similarity": float(row[4]) if row[4] is not None else 0.0,
                            "weighted_score": score
                        })
                
                if scores:
                    scores_array = np.array(scores)
                    
                    # 计算统计信息
                    agent_analysis = {
                        "name": sheet_name,
                        "total_questions": len(scores),
                        "mean_score": float(np.mean(scores_array)),
                        "median_score": float(np.median(scores_array)),
                        "max_score": float(np.max(scores_array)),
                        "min_score": float(np.min(scores_array)),
                        "std_score": float(np.std(scores_array)),
                        "high_score_ratio": float(np.sum(scores_array >= 0.8) / len(scores_array)),
                        "low_score_ratio": float(np.sum(scores_array < 0.5) / len(scores_array)),
                        "score_distribution": {
                            "excellent": int(np.sum(scores_array >= 0.9)),  # 优秀 (>=0.9)
                            "good": int(np.sum((scores_array >= 0.7) & (scores_array < 0.9))),  # 良好 (0.7-0.9)
                            "fair": int(np.sum((scores_array >= 0.5) & (scores_array < 0.7))),  # 一般 (0.5-0.7)
                            "poor": int(np.sum(scores_array < 0.5))  # 较差 (<0.5)
                        },
                        "questions": questions_data
                    }
                    
                    analysis_result["agents"][sheet_name] = agent_analysis
                    agent_scores[sheet_name] = np.mean(scores_array)
            
            # 确定最佳智能体
            if agent_scores:
                best_agent = max(agent_scores, key=agent_scores.get)
                analysis_result["best_agent"] = best_agent
                
                # 智能体比较
                sorted_agents = sorted(agent_scores.items(), key=lambda x: x[1], reverse=True)
                analysis_result["comparison"]["ranking"] = [
                    {"agent": agent, "score": float(score)} for agent, score in sorted_agents
                ]
                
                # 整体统计
                all_scores = []
                for agent_data in analysis_result["agents"].values():
                    all_scores.extend([q["weighted_score"] for q in agent_data["questions"]])
                
                if all_scores:
                    all_scores_array = np.array(all_scores)
                    analysis_result["overall_stats"] = {
                        "total_questions": len(all_scores),
                        "mean_score": float(np.mean(all_scores_array)),
                        "median_score": float(np.median(all_scores_array)),
                        "max_score": float(np.max(all_scores_array)),
                        "min_score": float(np.min(all_scores_array)),
                        "std_score": float(np.std(all_scores_array))
                    }
            
            return analysis_result
            
        except Exception as e:
            return {"error": f"分析文件失败: {str(e)}"}

    def get_question_details(self, file_path: str, agent_name: str = None) -> List[Dict]:
        """获取问题详细信息"""
        try:
            wb = load_workbook(file_path)
            
            if agent_name and agent_name in wb.sheetnames:
                sheets_to_process = [agent_name]
            else:
                sheets_to_process = wb.sheetnames
            
            questions = []
            
            for sheet_name in sheets_to_process:
                ws = wb[sheet_name]
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) >= 6 and row[0]:  # 确保有问题内容
                        question_data = {
                            "id": f"{sheet_name}_{row_idx}",
                            "agent": sheet_name,
                            "question": str(row[0]),
                            "standard_answer": str(row[1]) if row[1] else "",
                            "generated_answer": str(row[2]) if row[2] else "",
                            "cosine_similarity": float(row[3]) if row[3] is not None else 0.0,
                            "jaccard_similarity": float(row[4]) if row[4] is not None else 0.0,
                            "weighted_score": float(row[5]) if row[5] is not None else 0.0
                        }
                        questions.append(question_data)
            
            # 按综合评分排序
            questions.sort(key=lambda x: x["weighted_score"], reverse=True)
            return questions
            
        except Exception as e:
            return []

    def compare_agents(self, file_path: str) -> Dict[str, Any]:
        """比较不同智能体的表现"""
        try:
            wb = load_workbook(file_path)
            comparison = {
                "agents": [],
                "metrics": {
                    "mean_scores": {},
                    "high_score_ratios": {},
                    "low_score_ratios": {},
                    "consistency": {}  # 标准差，越小越一致
                }
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                scores = []
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 6 and row[5] is not None:
                        scores.append(float(row[5]))
                
                if scores:
                    scores_array = np.array(scores)
                    
                    agent_metrics = {
                        "name": sheet_name,
                        "mean_score": float(np.mean(scores_array)),
                        "high_score_ratio": float(np.sum(scores_array >= 0.8) / len(scores_array)),
                        "low_score_ratio": float(np.sum(scores_array < 0.5) / len(scores_array)),
                        "consistency": float(1 / (1 + np.std(scores_array)))  # 一致性指标
                    }
                    
                    comparison["agents"].append(agent_metrics)
                    comparison["metrics"]["mean_scores"][sheet_name] = agent_metrics["mean_score"]
                    comparison["metrics"]["high_score_ratios"][sheet_name] = agent_metrics["high_score_ratio"]
                    comparison["metrics"]["low_score_ratios"][sheet_name] = agent_metrics["low_score_ratio"]
                    comparison["metrics"]["consistency"][sheet_name] = agent_metrics["consistency"]
            
            # 排序
            comparison["agents"].sort(key=lambda x: x["mean_score"], reverse=True)
            
            return comparison
            
        except Exception as e:
            return {"error": f"比较智能体失败: {str(e)}"}

    def get_performance_insights(self, file_path: str) -> Dict[str, Any]:
        """获取性能洞察"""
        try:
            analysis = self.compare_agents(file_path)
            
            if "error" in analysis:
                return analysis
            
            insights = {
                "summary": {},
                "recommendations": [],
                "strengths": {},
                "weaknesses": {}
            }
            
            agents = analysis["agents"]
            
            if agents:
                best_agent = agents[0]
                worst_agent = agents[-1]
                
                insights["summary"] = {
                    "best_performer": best_agent["name"],
                    "worst_performer": worst_agent["name"],
                    "performance_gap": best_agent["mean_score"] - worst_agent["mean_score"],
                    "total_agents": len(agents)
                }
                
                # 生成建议
                for agent in agents:
                    if agent["low_score_ratio"] > 0.3:
                        insights["recommendations"].append(
                            f"{agent['name']} 有 {agent['low_score_ratio']:.1%} 的低分回答，建议优化模型或提示词"
                        )
                    
                    if agent["consistency"] < 0.5:
                        insights["recommendations"].append(
                            f"{agent['name']} 回答一致性较差，建议调整温度参数或优化训练数据"
                        )
                
                # 识别优势和劣势
                for agent in agents:
                    strengths = []
                    weaknesses = []
                    
                    if agent["mean_score"] > 0.8:
                        strengths.append("整体表现优秀")
                    if agent["high_score_ratio"] > 0.6:
                        strengths.append("高质量回答比例高")
                    if agent["consistency"] > 0.7:
                        strengths.append("回答一致性好")
                    
                    if agent["mean_score"] < 0.6:
                        weaknesses.append("整体表现需要改进")
                    if agent["low_score_ratio"] > 0.2:
                        weaknesses.append("低质量回答比例较高")
                    if agent["consistency"] < 0.5:
                        weaknesses.append("回答一致性差")
                    
                    insights["strengths"][agent["name"]] = strengths
                    insights["weaknesses"][agent["name"]] = weaknesses
            
            return insights
            
        except Exception as e:
            return {"error": f"生成性能洞察失败: {str(e)}"}