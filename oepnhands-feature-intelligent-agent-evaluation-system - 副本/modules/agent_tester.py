import aiohttp
import asyncio
import os
from openpyxl import Workbook, load_workbook
from typing import Dict, List, Any
import json

class AgentTester:
    def __init__(self):
        pass

    async def test_connection(self, url: str, key: str) -> bool:
        """测试智能体连接"""
        payload = {
            "inputs": {},
            "query": "测试连接",
            "response_mode": "blocking",
            "user": "test_user"
        }
        headers = {
            "Authorization": key,
            "Content-Type": "application/json"
        }
        
        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(url, json=payload, headers=headers, timeout=10) as response:
                    return response.status == 200
        except:
            return False

    async def call_dify(self, agent_name: str, agent_config: Dict, question: str, user_id: str = "eval_user") -> str:
        """调用Dify智能体"""
        payload = {
            "inputs": {},
            "query": question,
            "response_mode": "blocking",
            "user": user_id
        }
        headers = {
            "Authorization": agent_config["key"],
            "Content-Type": "application/json"
        }
        
        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(agent_config["url"], json=payload, headers=headers, timeout=60) as response:
                    response.raise_for_status()
                    result = await response.json()
                    return result.get("answer", "")
        except Exception as e:
            return f"Error calling {agent_name}: {e}"

    def read_questions_from_excel(self, file_path: str) -> List[Dict]:
        """从Excel文件读取问答对"""
        wb = load_workbook(file_path)
        ws = wb.active
        questions = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # 确保问题和答案都存在
                questions.append({
                    "question": str(row[0]),
                    "answer": str(row[1])
                })
        
        return questions

    def write_results(self, results_dict: Dict[str, List[Dict]], output_path: str):
        """将结果写入Excel文件"""
        wb = Workbook()
        # 删除默认sheet
        wb.remove(wb.active)

        for agent_name, rows in results_dict.items():
            ws = wb.create_sheet(title=agent_name)
            ws.append(["问题", "标准答案", f"{agent_name} 生成的答案"])
            
            for r in rows:
                ws.append([r["question"], r["standard_answer"], r["agent_answer"]])
                
            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min((max_length + 2), 50)
                ws.column_dimensions[column].width = adjusted_width
        
        wb.save(output_path)
        print(f"结果已保存至 {output_path}")

    async def test_agents(self, agents_config: Dict, file_paths: List[str], output_path: str, delay: int = 1) -> int:
        """测试智能体（从DOCX文件生成问答对）"""
        # 首先需要生成问答对
        from .qa_generator import QAGenerator
        
        # 这里需要API密钥，暂时使用默认值
        # 在实际使用中应该从配置中获取
        qa_generator = QAGenerator("sk-elnmwevbokezmyjvyafilsfsvdgwqbgsrjvlrnfhzsodtakc")
        temp_qa_file = f"temp_qa_{os.path.basename(output_path)}"
        
        questions = await qa_generator.process_documents(file_paths, temp_qa_file)
        
        # 然后测试智能体
        return await self.test_agents_with_qa_file(agents_config, temp_qa_file, output_path, delay)

    async def test_agents_with_qa_file(self, agents_config: Dict, qa_file_path: str, output_path: str, delay: int = 1) -> int:
        """使用已有的问答对文件测试智能体"""
        questions = self.read_questions_from_excel(qa_file_path)
        
        if not questions:
            raise Exception("没有找到有效的问答对")

        # 准备智能体配置
        active_agents = {}
        for i in range(1, agents_config["count"] + 1):
            agent_key = f"agent{i}"
            if agent_key in agents_config:
                active_agents[f"智能体{i}"] = agents_config[agent_key]

        results = {agent_name: [] for agent_name in active_agents.keys()}

        # 测试每个问题
        for idx, item in enumerate(questions, 1):
            print(f"正在处理第 {idx}/{len(questions)} 个问题...")
            question = item["question"]
            
            # 并发调用所有智能体
            tasks = []
            for agent_name, agent_config in active_agents.items():
                task = self.call_dify(agent_name, agent_config, question, f"{agent_name}_user_{idx}")
                tasks.append((agent_name, task))
            
            # 等待所有调用完成
            for agent_name, task in tasks:
                answer = await task
                results[agent_name].append({
                    "question": question,
                    "standard_answer": item["answer"],
                    "agent_answer": answer
                })
            
            # 控制请求频率
            if delay > 0:
                await asyncio.sleep(delay)

        self.write_results(results, output_path)
        return len(questions)