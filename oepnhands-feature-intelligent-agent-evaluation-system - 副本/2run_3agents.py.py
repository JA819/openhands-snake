import requests
import time
import os
from openpyxl import Workbook, load_workbook
from fastapi import FastAPI, File, UploadFile, Form, HTTPException, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse
from starlette.requests import Request

# --------------------------
# 1. 三个 Agent 的配置
# --------------------------
AGENTS = {
    "bge-m3": {
        "url": "https://aigov-dify.bytebroad.com.cn/v1/chat-messages",
        "key": "Bearer app-0R77MRBexBD7iCn87aoC1X69"
    },
    "qwen": {
        "url": "https://aigov-dify.bytebroad.com.cn/v1/chat-messages",
        "key": "Bearer app-TAwq1XsLmXXLzYKTQTJV2nUe"
    },
    "rangflow": {
        "url": "https://aigov-dify.bytebroad.com.cn/v1/chat-messages",
        "key": "Bearer app-AeGMfmHfQl2pMN5xVKYIbGEG"
    }
}

# --------------------------
# 2. 文件路径
# --------------------------
GOLDEN_SET_FILE = r"C:/Users/19541/Desktop/Internship services/Internship services/Dify/learnproject/learn HTTP/问答对1.xlsx"
OUTPUT_FILE     = r"C:/Users/19541/Desktop/Internship services/Internship services/Dify/learnproject/learn HTTP/evaluation_results_3agents.xlsx"

# --------------------------
# 3. 通用函数
# --------------------------
def call_dify(agent_name, question, user_id="eval_user"):
    cfg = AGENTS[agent_name]
    payload = {
        "inputs": {},
        "query": question,
        "response_mode": "blocking",
        "user": user_id
    }
    headers = {
        "Authorization": cfg["key"],
        "Content-Type": "application/json"
    }
    try:
        resp = requests.post(cfg["url"], json=payload, headers=headers, timeout=60)
        resp.raise_for_status()
        return resp.json().get("answer", "")
    except Exception as e:
        return f"Error calling {agent_name}: {e}"

def read_questions(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    return [{"question": r[0], "answer": r[1]}
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[0] and r[1]]

def write_results(results_dict):
    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    for agent, rows in results_dict.items():
        ws = wb.create_sheet(title=agent)
        ws.append(["问题", "标准答案", f"{agent} 生成的答案"])
        for r in rows:
            ws.append([r["question"], r["standard_answer"], r["agent_answer"]])
    wb.save(OUTPUT_FILE)
    print(f"结果已保存至 {OUTPUT_FILE}")

# --------------------------
# 4. 主流程
# --------------------------
if __name__ == "__main__":
    questions = read_questions(GOLDEN_SET_FILE)
    results = {k: [] for k in AGENTS}

    for idx, item in enumerate(questions, 1):
        print(f"正在处理第 {idx}/{len(questions)} 个问题 ...")
        q = item["question"]
        for agent in AGENTS:
            ans = call_dify(agent, q, f"{agent}_user_{idx}")
            results[agent].append({
                "question": q,
                "standard_answer": item["answer"],
                "agent_answer": ans
            })
        time.sleep(1.5)   # 控制频率

    write_results(results)
    print("全部完成！")

app = FastAPI()

# 挂载静态文件和模板
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# 创建必要的目录
os.makedirs("uploads", exist_ok=True)
os.makedirs("outputs", exist_ok=True)

# API路由
@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/api/files")
async def list_files():
    """获取文件列表"""
    files = []
    for file in os.listdir("outputs"):
        file_path = os.path.join("outputs", file)
        stats = os.stat(file_path)
        files.append({
            "name": file,
            "size": stats.st_size,
            "created_at": stats.st_ctime,
            "modified_at": stats.st_mtime,
            "type": get_file_type(file)
        })
    return files

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """文件上传处理"""
    try:
        file_path = os.path.join("uploads", file.filename)
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        return {"filename": file.filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ...其他API路由

const API = {
    async get(url) {
        const response = await fetch(url);
        if (!response.ok) {
            throw new Error(`HTTP error! status: ${response.status}`);
        }
        return await response.json();
    },
    
    async post(url, data) {
        const response = await fetch(url, {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
            },
            body: JSON.stringify(data)
        });
        if (!response.ok) {
            throw new Error(`HTTP error! status: ${response.status}`);
        }
        return await response.json();
    },
    
    async postForm(url, formData) {
        const response = await fetch(url, {
            method: 'POST',
            body: formData
        });
        if (!response.ok) {
            throw new Error(`HTTP error! status: ${response.status}`);
        }
        return await response.json();
    }
};

const Message = {
    show(message, type = 'info') {
        // 实现消息提示
        toastr[type](message);
    },
    success(message) {
        this.show(message, 'success');
    },
    error(message) {
        this.show(message, 'error');
    }
};

const Utils = {
    formatDate(timestamp) {
        return new Date(timestamp * 1000).toLocaleString();
    },
    
    formatFileSize(bytes) {
        const units = ['B', 'KB', 'MB', 'GB'];
        let size = bytes;
        let unitIndex = 0;
        while (size >= 1024 && unitIndex < units.length - 1) {
            size /= 1024;
            unitIndex++;
        }
        return `${size.toFixed(2)} ${units[unitIndex]}`;
    }
};
