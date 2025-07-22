import requests, json, os, jieba, numpy as np
from openpyxl import load_workbook
from sklearn.metrics.pairwise import cosine_similarity

# --------------------------
# 1. 配置
# --------------------------
EMBED_URL = "https://api.siliconflow.cn/v1/embeddings"
API_KEY   = "Bearer sk-elnmwevbokezmyjvyafilsfsvdgwqbgsrjvlrnfhzsodtakc"

INPUT_FILE  = r"C:/Users/19541/Desktop/Internship services/Internship services/Dify/learnproject/learn HTTP/evaluation_results_3agents.xlsx"
OUTPUT_FILE = r"C:/Users/19541\Desktop/Internship services/Internship services\Dify/learnproject/learn HTTP/evaluation_results_3agents_with_scores.xlsx"

HEADERS = {
    "Authorization": API_KEY,
    "Content-Type": "application/json"
}

# --------------------------
# 2. 获取向量
# --------------------------
def get_embedding(text: str, retries: int = 3) -> list:
    if not isinstance(text, str) or not text.strip():
        return []
    payload = {
        "model": "Qwen/Qwen3-Embedding-0.6B",
        "input": " ".join(jieba.lcut(text)),
        "encoding_format": "float"
    }
    for _ in range(retries):
        try:
            r = requests.post(EMBED_URL, json=payload, headers=HEADERS, timeout=30)
            r.raise_for_status()
            emb = r.json()["data"][0]["embedding"]
            return emb
        except Exception as e:
            print(f"embedding 失败: {e}，重试...")
    return []

# --------------------------
# 3. 相似度计算
# --------------------------
def similarity_scores(std: str, gen: str):
    emb_std = get_embedding(std)
    emb_gen = get_embedding(gen)
    cos = 0.0
    if emb_std and emb_gen and len(emb_std) == len(emb_gen):
        cos = cosine_similarity([emb_std], [emb_gen])[0][0]

    set_std = set(jieba.lcut(std))
    set_gen = set(jieba.lcut(gen))
    jac = len(set_std & set_gen) / len(set_std | set_gen) if set_std | set_gen else 0.0

    weighted = 0.7 * cos + 0.3 * jac
    return cos, jac, weighted

# --------------------------
# 4. 处理每个 sheet
# --------------------------
def main():
    wb = load_workbook(INPUT_FILE)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 加表头（仅当不存在时才加）
        if ws.cell(row=1, column=4).value != "余弦相似度":
            ws["D1"] = "余弦相似度"
            ws["E1"] = "Jaccard相似度"
            ws["F1"] = "综合相似度评分"
        # 从第2行开始遍历
        for row in ws.iter_rows(min_row=2, max_col=6, values_only=False):
            std, gen = row[1].value, row[2].value
            if std and gen:
                cos, jac, w = similarity_scores(str(std), str(gen))
                row[3].value, row[4].value, row[5].value = cos, jac, w
    wb.save(OUTPUT_FILE)
    print("全部 sheet 处理完成！结果已保存至", OUTPUT_FILE)

if __name__ == "__main__":
    main()
